#!/usr/bin/env python3
"""
GSCMTF Refresh Script
=====================
Reads data from GSCMTF_Input_Tracker_v2.xlsx and injects it as JSON
into gscmtf_dashboard.html, updating the dashboard automatically.

Usage:
    python gscmtf_refresh.py

Both files must be in the same folder as this script.

Requirements:
    pip install openpyxl
"""

import json
import re
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

# ── File paths (same folder as this script) ───────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE   = os.path.join(SCRIPT_DIR, "GSCMTF_Input_Tracker_v2.xlsx")
HTML_FILE    = os.path.join(SCRIPT_DIR, "gscmtf_dashboard.html")

# ── Country sheet names and IDs ───────────────────────────────────────────────
COUNTRIES = [
    ("BGD", "bangladesh", "Bangladesh", "🇧🇩", "Asia"),
    ("TCD", "chad",       "Chad",       "🇹🇩", "West Africa"),
    ("COL", "colombia",   "Colombia",   "🇨🇴", "Latin America"),
    ("DRC", "drc",        "DRC",        "🇨🇩", "Central Africa"),
    ("ETH", "ethiopia",   "Ethiopia",   "🇪🇹", "East Africa"),
    ("HTI", "haiti",      "Haiti",      "🇭🇹", "Caribbean"),
    ("HND", "honduras",   "Honduras",   "🇭🇳", "Central America"),
    ("MMR", "myanmar",    "Myanmar",    "🇲🇲", "South-East Asia"),
    ("SYR", "syria",      "Syria",      "🇸🇾", "Middle East"),
]

# ── Fixed row positions (identical across all 9 country sheets) ───────────────
# Section 1 – Programmatic
ROW_STATUS          = 6
ROW_BUDGET_TOTAL    = 7
ROW_BUDGET_SPENT    = 8   # computed via SUMIF on Budget Detail
ROW_BUDGET_COMMITTED= 9   # computed via SUMIF on Budget Detail
ROW_BENE_TARGET     = 12
ROW_BENE_REACHED    = 13
ROW_DIST_TOTAL      = 15
ROW_DIST_COMPLETED  = 16
ROW_NEXT_DIST       = 18
ROW_START_DATE      = 19
ROW_END_DATE        = 20
ROW_OFFICER         = 21
ROW_FAO_OFFICE      = 22

# Section 2 – Commodities (B25:B32)
ROW_ITEMS_START     = 25
ROW_ITEMS_END       = 32

# Section 3 – Orders pipeline
ROW_PENDING         = 35  # COUNTIF on Orders Detail
ROW_CONFIRMED       = 36
ROW_IN_PRODUCTION   = 37
ROW_SHIPPED         = 38
ROW_DELIVERED       = 39
ROW_OFR             = 41  # Order Fulfilment Rate (manual %)

# Section 4 – Inventory
ROW_ON_HAND         = 44
ROW_IN_TRANSIT      = 45
ROW_CAPACITY        = 47
ROW_UNIT            = 48
ROW_LOC1            = 49
ROW_LOC2            = 50
ROW_LOC3            = 51
ROW_LOC4            = 52

# Section 5 – Deliveries
ROW_PLANNED         = 55
ROW_DISPATCHED      = 56
ROW_RECV_PARTNERS   = 57
ROW_LAST_MILE       = 58
ROW_TRANSPORT       = 62
ROW_PARTNER1        = 63
ROW_PARTNER2        = 64
ROW_PARTNER3        = 65
ROW_PARTNER4        = 66

# Section 6 – KPIs
ROW_OTD             = 69
ROW_STOCKOUT        = 70
ROW_ASSESSMENT      = 71

# Section 7 – Risks (rows 76–83, cols B=Level, C=Description, D=Mitigation)
ROW_RISKS_START     = 76
ROW_RISKS_END       = 83

# Section 8 – Budget Detail (SUMIF source)
ROW_BD_DATA_START   = 91
ROW_BD_DATA_END     = 125
COL_BD_AMOUNT       = 6   # Column F
COL_BD_STATUS       = 8   # Column H  ("Spent" / "Committed" / "Planned")

# Section 9 – Orders Detail (COUNTIF source)
ROW_OD_DATA_START   = 129
ROW_OD_DATA_END     = 178
COL_OD_STATUS       = 13  # Column M


# ── Helper functions ──────────────────────────────────────────────────────────
def txt(v):
    """Return stripped string or empty string."""
    if v is None:
        return ""
    return str(v).strip()

def num(v, default=0):
    """Return float or default."""
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default

def pct(v):
    """Return percentage value (0–100). Handles both 0.75 and 75 formats."""
    n = num(v, 0)
    if 0 < n <= 1:
        return round(n * 100, 1)
    return round(n, 1)

def sumif(ws, criteria_col, criteria_val, sum_col, row_start, row_end):
    """Python-side SUMIF — avoids dependency on cached Excel formula values."""
    total = 0.0
    for r in range(row_start, row_end + 1):
        crit = txt(ws.cell(row=r, column=criteria_col).value)
        if crit.lower() == criteria_val.lower():
            total += num(ws.cell(row=r, column=sum_col).value, 0)
    return total

def countif(ws, col, criteria_val, row_start, row_end):
    """Python-side COUNTIF."""
    count = 0
    for r in range(row_start, row_end + 1):
        v = txt(ws.cell(row=r, column=col).value)
        if v.lower() == criteria_val.lower():
            count += 1
    return count

def cell(ws, row, col=2):
    """Read cell value by row (default col B=2)."""
    return ws.cell(row=row, column=col).value


# ── Read one country sheet ────────────────────────────────────────────────────
def read_country(wb, sheet_id, country_id, country_name, flag, region):
    if sheet_id not in wb.sheetnames:
        print(f"  WARNING: sheet '{sheet_id}' not found — skipping")
        return None

    ws = wb[sheet_id]

    # -- Programmatic --
    status_raw = txt(cell(ws, ROW_STATUS)).lower()
    if "on" in status_raw and "track" in status_raw:
        status = "on-track"
    elif "risk" in status_raw or "delay" in status_raw:
        status = "at-risk"
    elif "critical" in status_raw or "stop" in status_raw:
        status = "critical"
    else:
        status = "not-started"

    budget_total     = num(cell(ws, ROW_BUDGET_TOTAL))
    budget_spent     = sumif(ws, COL_BD_STATUS, "Spent",     COL_BD_AMOUNT, ROW_BD_DATA_START, ROW_BD_DATA_END)
    budget_committed = sumif(ws, COL_BD_STATUS, "Committed", COL_BD_AMOUNT, ROW_BD_DATA_START, ROW_BD_DATA_END)

    # Fall back to manually entered values if Budget Detail is empty
    if budget_spent == 0:
        budget_spent = num(cell(ws, ROW_BUDGET_SPENT))
    if budget_committed == 0:
        budget_committed = num(cell(ws, ROW_BUDGET_COMMITTED))

    bene_target   = int(num(cell(ws, ROW_BENE_TARGET)))
    bene_reached  = int(num(cell(ws, ROW_BENE_REACHED)))
    dist_total    = int(num(cell(ws, ROW_DIST_TOTAL)))
    dist_completed= int(num(cell(ws, ROW_DIST_COMPLETED)))
    next_dist     = txt(cell(ws, ROW_NEXT_DIST)) or "TBC"
    start_date    = txt(cell(ws, ROW_START_DATE)) or "TBC"
    end_date      = txt(cell(ws, ROW_END_DATE))   or "TBC"
    officer       = txt(cell(ws, ROW_OFFICER))
    fao_office    = txt(cell(ws, ROW_FAO_OFFICE))

    # Commodities
    items = []
    for r in range(ROW_ITEMS_START, ROW_ITEMS_END + 1):
        v = txt(cell(ws, r))
        if v and v != "—":
            items.append(v)
    if not items:
        items = ["—"]

    # -- Orders --
    pending       = countif(ws, COL_OD_STATUS, "Pending",       ROW_OD_DATA_START, ROW_OD_DATA_END)
    confirmed     = countif(ws, COL_OD_STATUS, "Confirmed",     ROW_OD_DATA_START, ROW_OD_DATA_END)
    in_production = countif(ws, COL_OD_STATUS, "In Production", ROW_OD_DATA_START, ROW_OD_DATA_END)
    shipped       = countif(ws, COL_OD_STATUS, "Shipped",       ROW_OD_DATA_START, ROW_OD_DATA_END)
    delivered     = countif(ws, COL_OD_STATUS, "Delivered",     ROW_OD_DATA_START, ROW_OD_DATA_END)

    # Fall back to manually entered order counts if Orders Detail is empty
    if (pending + confirmed + in_production + shipped + delivered) == 0:
        pending       = int(num(cell(ws, ROW_PENDING)))
        confirmed     = int(num(cell(ws, ROW_CONFIRMED)))
        in_production = int(num(cell(ws, ROW_IN_PRODUCTION)))
        shipped       = int(num(cell(ws, ROW_SHIPPED)))
        delivered     = int(num(cell(ws, ROW_DELIVERED)))

    # -- Inventory --
    on_hand    = num(cell(ws, ROW_ON_HAND))
    in_transit = num(cell(ws, ROW_IN_TRANSIT))
    capacity   = num(cell(ws, ROW_CAPACITY), 1000)
    unit       = txt(cell(ws, ROW_UNIT)) or "MT"
    locations  = [txt(cell(ws, r)) for r in range(ROW_LOC1, ROW_LOC4 + 1)
                  if txt(cell(ws, r)) and txt(cell(ws, r)) != "—"]
    if not locations:
        locations = ["—"]

    # -- Deliveries --
    planned        = int(num(cell(ws, ROW_PLANNED)))
    dispatched     = int(num(cell(ws, ROW_DISPATCHED)))
    recv_partners  = int(num(cell(ws, ROW_RECV_PARTNERS)))
    last_mile      = int(num(cell(ws, ROW_LAST_MILE)))
    transport      = txt(cell(ws, ROW_TRANSPORT)) or "—"
    partners       = [txt(cell(ws, r)) for r in range(ROW_PARTNER1, ROW_PARTNER4 + 1)
                      if txt(cell(ws, r)) and txt(cell(ws, r)) != "—"]
    if not partners:
        partners = ["—"]

    # -- KPIs --
    otd_raw      = cell(ws, ROW_OTD)
    stockout_raw = txt(cell(ws, ROW_STOCKOUT)) or "—"
    ofr_raw      = cell(ws, ROW_OFR)

    on_time_delivery  = pct(otd_raw)
    order_fulfillment = pct(ofr_raw)

    # Normalise stockout risk label
    sr = stockout_raw.lower()
    if "high" in sr:
        stockout_risk = "High"
    elif "med" in sr:
        stockout_risk = "Medium"
    elif "low" in sr:
        stockout_risk = "Low"
    else:
        stockout_risk = stockout_raw or "—"

    # -- Risks --
    risks = []
    for r in range(ROW_RISKS_START, ROW_RISKS_END + 1):
        level = txt(ws.cell(row=r, column=2).value).lower()
        desc  = txt(ws.cell(row=r, column=3).value)
        mitig = txt(ws.cell(row=r, column=4).value)
        if desc:
            risks.append({
                "level":      level if level in ("high", "medium", "low") else "low",
                "desc":       desc,
                "mitigation": mitig,
                "owner":      ""
            })

    # -- Assemble country object --
    return {
        "id":     country_id,
        "name":   country_name,
        "flag":   flag,
        "region": region,
        "status": status,
        "programmatic": {
            "budget": {
                "total":     budget_total,
                "spent":     round(budget_spent, 2),
                "committed": round(budget_committed, 2)
            },
            "beneficiaries": {
                "target":  bene_target,
                "reached": bene_reached
            },
            "distributions": {
                "totalMonths":     dist_total,
                "completedMonths": dist_completed,
                "nextDist":        next_dist
            },
            "items":         items,
            "startDate":     start_date,
            "endDate":       end_date,
            "projectOfficer":officer,
            "faoOffice":     fao_office
        },
        "sc": {
            "orders": {
                "pending":      pending,
                "confirmed":    confirmed,
                "inProduction": in_production,
                "shipped":      shipped,
                "delivered":    delivered
            },
            "inventory": {
                "onHand":    on_hand,
                "inTransit": in_transit,
                "capacity":  capacity,
                "unit":      unit,
                "locations": locations
            },
            "deliveries": {
                "planned":           planned,
                "dispatched":        dispatched,
                "receivedByPartners":recv_partners,
                "lastMile":          last_mile
            },
            "partners":  partners,
            "transport": transport,
            "kpis": {
                "orderFulfillment": order_fulfillment,
                "onTimeDelivery":   on_time_delivery,
                "stockoutRisk":     stockout_risk
            },
            "risks": risks
        }
    }


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  GSCMTF Dashboard Refresh")
    print("=" * 60)

    # Check files exist
    if not os.path.exists(EXCEL_FILE):
        print(f"\nERROR: Excel file not found:\n  {EXCEL_FILE}")
        print("Make sure GSCMTF_Input_Tracker_v2.xlsx is in the same folder.")
        return

    if not os.path.exists(HTML_FILE):
        print(f"\nERROR: HTML file not found:\n  {HTML_FILE}")
        print("Make sure gscmtf_dashboard.html is in the same folder.")
        return

    # Load workbook
    print(f"\nReading: {os.path.basename(EXCEL_FILE)}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    print(f"Sheets found: {', '.join(wb.sheetnames)}")

    # Read all countries
    print("\nProcessing countries:")
    all_countries = []
    for sheet_id, country_id, name, flag, region in COUNTRIES:
        print(f"  {flag}  {name} ({sheet_id}) ... ", end="")
        data = read_country(wb, sheet_id, country_id, name, flag, region)
        if data:
            all_countries.append(data)
            print(f"OK  [budget: ${data['programmatic']['budget']['total']:,.0f} | "
                  f"status: {data['status']}]")
        else:
            print("SKIPPED")

    if not all_countries:
        print("\nERROR: No country data could be read.")
        return

    # Build JSON block
    data_json  = json.dumps(all_countries, indent=2, ensure_ascii=False)
    new_block  = (
        "// ==GSCMTF_DATA_START==\n"
        "const countries = \n"
        + data_json + ";\n"
        "// ==GSCMTF_DATA_END=="
    )

    # Read HTML
    print(f"\nReading: {os.path.basename(HTML_FILE)}")
    with open(HTML_FILE, "r", encoding="utf-8") as f:
        html = f.read()

    # Check markers exist
    if "==GSCMTF_DATA_START==" not in html:
        print("ERROR: Data injection markers not found in HTML file.")
        print("Make sure gscmtf_dashboard.html is the correct version.")
        return

    # Inject data
    pattern  = r"// ==GSCMTF_DATA_START==.*?// ==GSCMTF_DATA_END=="
    new_html = re.sub(pattern, new_block, html, flags=re.DOTALL)

    # Update "Last updated" timestamp
    ts = datetime.now().strftime("%d %b %Y %H:%M")
    new_html = re.sub(
        r'id="last-updated">[^<]*<',
        f'id="last-updated">Last updated: {ts}<',
        new_html
    )

    # Write back
    with open(HTML_FILE, "w", encoding="utf-8") as f:
        f.write(new_html)

    print(f"\n✓ Dashboard updated successfully ({len(all_countries)} countries)")
    print(f"✓ Timestamp: {ts}")
    print(f"✓ File: {HTML_FILE}")
    print("\nOpen gscmtf_dashboard.html in your browser to view the result.")
    print("=" * 60)


if __name__ == "__main__":
    main()
